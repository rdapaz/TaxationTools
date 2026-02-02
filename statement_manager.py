import re
import sys
import sqlite3
import argparse
from decimal import Decimal
from datetime import datetime
from pathlib import Path
import pyperclip
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


class StatementManager:
    def __init__(self, db_path, config_path=None):
        self.db_path = db_path
        self.conn = None
        self.categories = [
            "software",
            "professional membership",
            "technical library",
            "magazines and journals",
            "training",
            "not work related",
            "other"
        ]
        self.api_key = None
        self.client = None

        if config_path:
            self._load_config(config_path)

        self._init_database()

    def _load_config(self, config_path):
        """Load configuration from JSON file"""
        try:
            import json
            with open(config_path, 'r') as f:
                config = json.load(f)
                self.api_key = config.get('anthropic_api_key')
                if self.api_key:
                    import anthropic
                    self.client = anthropic.Anthropic(api_key=self.api_key)
                    print("✓ Claude API configured")
        except Exception as e:
            print(f"Warning: Could not load config file: {e}")

    def _init_database(self):
        """Initialize database and create table if it doesn't exist"""
        self.conn = sqlite3.connect(self.db_path)
        cursor = self.conn.cursor()
        cursor.execute("""
                       CREATE TABLE IF NOT EXISTS transactions
                       (
                           id               INTEGER PRIMARY KEY AUTOINCREMENT,
                           trans_date       TEXT NOT NULL,
                           description      TEXT NOT NULL,
                           reference_number TEXT NOT NULL,
                           amount           REAL NOT NULL,
                           category         TEXT,
                           ai_classified    INTEGER   DEFAULT 0,
                           created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                           UNIQUE (trans_date, reference_number, amount)
                       )
                       """)
        self.conn.commit()

    def parse_clipboard(self):
        """Parse clipboard data and extract transactions - cross-platform"""
        try:
            data = pyperclip.paste()
        except Exception as e:
            print(f"Error accessing clipboard: {e}")
            print("Make sure you have clipboard access enabled for this application.")
            return []

        if not data:
            print("Clipboard is empty")
            return []

        transactions = []
        lines = data.splitlines()

        # Pattern to match date at start of line
        date_pattern = re.compile(r'^(\d{6})\s+(.+)')
        # Pattern to match reference number (long digit sequence)
        ref_pattern = re.compile(r'\d{15,}')

        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue

            match = date_pattern.match(line)
            if match:
                date_str = match.group(1)
                rest = match.group(2).strip()

                # Check if amount is on this line or next
                amount_match = re.search(r'([\d.]+)$', rest)

                if amount_match:
                    # Amount is on same line
                    amount = Decimal(amount_match.group(1))
                    desc_and_ref = rest[:amount_match.start()].strip()

                    # Extract reference number from next line if available
                    ref_num = ""
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        ref_match = ref_pattern.search(next_line)
                        if ref_match:
                            ref_num = ref_match.group(0)
                            i += 1  # Skip the reference line

                    # Format date as YYYY-MM-DD
                    formatted_date = f"20{date_str[4:6]}-{date_str[2:4]}-{date_str[0:2]}"

                    transactions.append({
                        'date': formatted_date,
                        'description': desc_and_ref,
                        'reference': ref_num,
                        'amount': float(amount)
                    })

            i += 1

        return transactions

    def classify_transaction(self, description: str, amount: float):
        """Use Claude API to classify a transaction"""
        if not self.client:
            return None

        try:
            prompt = f"""Given this credit card transaction, classify it into one of these categories for tax deduction purposes:

Categories:
- "software": Software purchases, SaaS subscriptions, development tools
- "professional membership": Professional organization memberships, certifications
- "technical library": Books, technical publications, O'Reilly, Manning, etc.
- "magazines and journals": Technical magazines, IEEE, ACM subscriptions
- "training": Online courses, conferences, workshops, training materials
- "not work related": Personal purchases, food, entertainment, general shopping
- "other": Work-related but doesn't fit other categories

Transaction:
Description: {description}
Amount: ${amount}

Context: This is for an OT cybersecurity professional in industrial control systems.

Respond with ONLY the category name in lowercase, nothing else."""

            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=50,
                messages=[{"role": "user", "content": prompt}]
            )

            category = message.content[0].text.strip().strip('"').lower()

            # Validate category
            if category in self.categories:
                return category
            else:
                return "other"

        except Exception as e:
            print(f"  Warning: Classification failed for '{description[:40]}...': {e}")
            return None

    def add_transactions(self, transactions, use_ai=False):
        """Add transactions to database, checking for duplicates"""
        cursor = self.conn.cursor()
        added = 0
        duplicates = []

        for i, trans in enumerate(transactions, 1):
            category = None
            ai_classified = 0

            # Try AI classification if enabled
            if use_ai and self.client:
                print(f"  Classifying {i}/{len(transactions)}: {trans['description'][:50]}...")
                category = self.classify_transaction(trans['description'], trans['amount'])
                if category:
                    ai_classified = 1
                    print(f"    → {category}")

            try:
                cursor.execute("""
                               INSERT INTO transactions (trans_date, description, reference_number, amount, category,
                                                         ai_classified)
                               VALUES (?, ?, ?, ?, ?, ?)
                               """, (trans['date'], trans['description'], trans['reference'], trans['amount'], category,
                                     ai_classified))
                added += 1
            except sqlite3.IntegrityError:
                duplicates.append(trans)

        self.conn.commit()

        return added, duplicates

    def export_to_excel(self, output_path):
        """Export transactions to Excel with monthly worksheets"""
        cursor = self.conn.cursor()
        cursor.execute("""
                       SELECT id, trans_date, description, reference_number, amount, category, ai_classified
                       FROM transactions
                       ORDER BY trans_date
                       """)

        transactions = cursor.fetchall()

        if not transactions:
            print("No transactions to export")
            return

        # Group by month
        monthly_data = {}
        for trans in transactions:
            trans_id = trans[0]
            date_obj = datetime.strptime(trans[1], '%Y-%m-%d')
            month_key = date_obj.strftime('%Y-%m')
            month_name = date_obj.strftime('%b %Y')

            if month_name not in monthly_data:
                monthly_data[month_name] = []

            monthly_data[month_name].append(trans)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create data validation for categories
        dv = DataValidation(type="list", formula1=f'"{",".join(self.categories)}"', allow_blank=True)

        for month_name, trans_list in sorted(monthly_data.items()):
            # Create worksheet (sanitize name for Excel)
            ws_name = month_name[:31]  # Excel limit
            ws = wb.create_sheet(title=ws_name)

            # Headers - Added ID column
            headers = ['ID', 'Date', 'Description', 'Reference', 'Amount', 'Category']
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for trans in trans_list:
                trans_id = trans[0]
                date_obj = datetime.strptime(trans[1], '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
                category = trans[5] or ''

                # Add indicator if AI classified
                if trans[6] == 1 and category:  # ai_classified flag
                    category = f"{category} 🤖"

                ws.append([trans_id, formatted_date, trans[2], trans[3], trans[4], category])

            # Add data validation to category column (now column F)
            last_row = ws.max_row
            dv.add(f'F2:F{last_row}')
            ws.add_data_validation(dv)

            # Adjust column widths
            ws.column_dimensions['A'].width = 8  # ID
            ws.column_dimensions['B'].width = 12  # Date
            ws.column_dimensions['C'].width = 45  # Description
            ws.column_dimensions['D'].width = 20  # Reference
            ws.column_dimensions['E'].width = 12  # Amount
            ws.column_dimensions['F'].width = 28  # Category

            # Format amount column as currency
            for row in range(2, last_row + 1):
                ws[f'E{row}'].number_format = '$#,##0.00'

        # Save workbook
        wb.save(output_path)
        print(f"\n✓ Excel file created: {output_path}")
        print(f"  Total worksheets: {len(monthly_data)}")

        # Count AI classifications
        ai_count = sum(1 for trans in transactions if trans[6] == 1)
        if ai_count > 0:
            print(f"  AI-classified entries: {ai_count} (marked with 🤖)")

    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()


def main():
    parser = argparse.ArgumentParser(
        description='Credit Card Statement Manager with AI Classification (Cross-platform)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Add transactions from clipboard
  python statement_manager.py -d statements.db -a

  # Add with AI classification
  python statement_manager.py -d statements.db -c config.json -a --classify

  # Export to Excel
  python statement_manager.py -d statements.db -e expenses.xlsx

  # Add with AI and export
  python statement_manager.py -d statements.db -c config.json -a --classify -e expenses.xlsx
        """
    )

    parser.add_argument('-d', '--database', required=True,
                        help='Path to SQLite database file')
    parser.add_argument('-c', '--config',
                        help='Path to config file (JSON with anthropic_api_key)')
    parser.add_argument('-a', '--add', action='store_true',
                        help='Add transactions from clipboard to database')
    parser.add_argument('--classify', action='store_true',
                        help='Use AI to classify transactions (requires -c and -a)')
    parser.add_argument('-e', '--export',
                        help='Export database to Excel file')

    args = parser.parse_args()

    # Validate arguments
    if args.classify and not args.add:
        print("Error: --classify requires -a (add) flag")
        sys.exit(1)

    if args.classify and not args.config:
        print("Error: --classify requires -c (config) flag with API key")
        sys.exit(1)

    # Initialize manager
    manager = StatementManager(args.database, args.config)

    try:
        # Add transactions if requested
        if args.add:
            print("Parsing clipboard data...")
            transactions = manager.parse_clipboard()

            if not transactions:
                print("No valid transactions found in clipboard")
            else:
                print(f"Found {len(transactions)} transactions\n")

                if args.classify:
                    print("AI Classification enabled...")

                added, duplicates = manager.add_transactions(transactions, use_ai=args.classify)

                print(f"\n✓ Added {added} new transactions")

                if duplicates:
                    print(f"\n⚠️  Found {len(duplicates)} duplicate(s):")
                    for dup in duplicates:
                        print(f"  {dup['date']} - {dup['description'][:50]} - ${dup['amount']}")
                    print("\nDuplicate transactions were not added to the database.")
                    sys.exit(1)

        # Export to Excel if requested
        if args.export:
            print(f"\nExporting to Excel...")
            manager.export_to_excel(args.export)

    finally:
        manager.close()


if __name__ == "__main__":
    main()