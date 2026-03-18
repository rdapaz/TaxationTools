"""
Transactions view — searchable, filterable list with inline category editing.
"""

from datetime import date, timedelta
from pathlib import Path

from PySide6.QtCore import Qt, Signal, QSortFilterProxyModel
from PySide6.QtGui import QColor, QShortcut, QKeySequence
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableView,
    QComboBox, QLineEdit, QDateEdit, QPushButton, QHeaderView,
    QFrame, QAbstractItemView, QStyledItemDelegate, QFileDialog,
    QMessageBox,
)
from PySide6.QtCore import QAbstractTableModel, QModelIndex, QDate

from gui.models.database import BudgetDatabase, Transaction
from gui.categories import CATEGORIES, get_groups, colour_for
from gui.widgets.help_window import make_help_button
from gui.theme import (
    font_heading, font_body, BG_MAIN, BG_CARD, BORDER, NAVY,
    TEXT_PRIMARY, TEXT_SECONDARY, CARD_STYLE,
    INPUT_STYLE, COMBO_STYLE, DATE_STYLE, TABLE_STYLE,
)


class CategoryDelegate(QStyledItemDelegate):
    """Dropdown delegate for editing the category column in-place."""

    def __init__(self, db: BudgetDatabase, parent=None):
        super().__init__(parent)
        self.db = db

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.setStyleSheet(f"""
            QComboBox {{
                font-size: 10pt;
                background-color: white;
                color: {TEXT_PRIMARY};
                padding: 2px 6px;
            }}
            QComboBox QAbstractItemView {{
                background-color: white;
                color: {TEXT_PRIMARY};
                border: 1px solid {BORDER};
                selection-background-color: #E8EDF5;
                selection-color: {NAVY};
                font-size: 10pt;
            }}
        """)
        for group_name, cats in get_groups():
            for cat in cats:
                combo.addItem(cat.display_name, cat.id)
        return combo

    def setEditorData(self, editor, index):
        txn = index.data(Qt.ItemDataRole.UserRole)
        if txn:
            idx = editor.findData(txn.category)
            if idx >= 0:
                editor.setCurrentIndex(idx)

    def setModelData(self, editor, model, index):
        new_cat_id = editor.currentData()
        txn = index.data(Qt.ItemDataRole.UserRole)
        if txn and new_cat_id and new_cat_id != txn.category:
            self.db.update_category(txn.id, new_cat_id)
            # Update in-memory
            txn.category = new_cat_id
            # Emit change for entire row so all columns repaint
            left = model.index(index.row(), 0)
            right = model.index(index.row(), model.columnCount() - 1)
            model.dataChanged.emit(left, right)


class TransactionTableModel(QAbstractTableModel):
    """Model backing the transaction table."""

    HEADERS = ["Date", "Description", "Amount", "Category"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._data: list[Transaction] = []

    def set_data(self, txns: list[Transaction]):
        self.beginResetModel()
        self._data = txns
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        return len(self.HEADERS)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self.HEADERS[section]
        return None

    def flags(self, index):
        base = super().flags(index)
        if index.column() == 3:  # Category column is editable
            return base | Qt.ItemFlag.ItemIsEditable
        return base

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        txn = self._data[index.row()]
        col = index.column()

        if role == Qt.ItemDataRole.DisplayRole:
            if col == 0:
                return txn.trans_date.strftime("%d/%m/%Y")
            elif col == 1:
                return txn.description
            elif col == 2:
                return f"${txn.amount:,.2f}"
            elif col == 3:
                cat = CATEGORIES.get(txn.category)
                return cat.display_name if cat else (txn.category or "Uncategorised").replace("_", " ").title()

        elif role == Qt.ItemDataRole.TextAlignmentRole:
            if col == 2:
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter

        elif role == Qt.ItemDataRole.ForegroundRole:
            if col == 3:
                return QColor(colour_for(txn.category or "uncategorised"))
            return QColor(TEXT_PRIMARY)

        elif role == Qt.ItemDataRole.UserRole:
            return txn

        return None

    def get_transaction(self, row: int) -> Transaction | None:
        if 0 <= row < len(self._data):
            return self._data[row]
        return None


class TransactionsView(QWidget):
    """Filterable transaction list with inline category editing."""

    def __init__(self, db: BudgetDatabase, parent=None):
        super().__init__(parent)
        self.db = db
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(16)
        self.setStyleSheet(f"QWidget {{ background: {BG_MAIN}; color: {TEXT_PRIMARY}; }}")

        # Title row with help button
        title_row = QHBoxLayout()
        title = QLabel("Transactions")
        title.setFont(font_heading(20))
        title.setStyleSheet(f"color: {NAVY}; background: transparent;")
        title_row.addWidget(title)
        title_row.addWidget(make_help_button("Transactions", self))
        title_row.addStretch()
        layout.addLayout(title_row)

        # Filter bar
        filter_card = QFrame()
        filter_card.setObjectName("card")
        filter_card.setStyleSheet(CARD_STYLE)
        filter_layout = QHBoxLayout(filter_card)
        filter_layout.setContentsMargins(16, 10, 16, 10)
        filter_layout.setSpacing(12)

        # Search
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search description or date (e.g. 2026-02, woolworths)...")
        self.search_input.setMinimumWidth(200)
        self.search_input.setStyleSheet(INPUT_STYLE)
        self.search_input.textChanged.connect(self._on_filter_changed)
        self.search_input.returnPressed.connect(self._on_filter_changed)
        filter_layout.addWidget(self.search_input, stretch=2)

        # Category filter
        self.cat_combo = QComboBox()
        self.cat_combo.addItem("All Categories", "")
        for group_name, cats in get_groups():
            for cat in cats:
                self.cat_combo.addItem(f"  {cat.display_name}", cat.id)
        self.cat_combo.setMinimumWidth(180)
        self.cat_combo.setStyleSheet(COMBO_STYLE)
        self.cat_combo.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.cat_combo)

        # Date range — default to full data range from DB
        min_date, max_date = self.db.get_date_range()

        lbl_from = QLabel("From:")
        lbl_from.setStyleSheet(f"color: {NAVY}; background: transparent; border: none;")
        filter_layout.addWidget(lbl_from)

        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd/MM/yyyy")
        if min_date:
            self.date_from.setDate(QDate.fromString(min_date, "yyyy-MM-dd"))
        else:
            self.date_from.setDate(QDate.currentDate().addYears(-1))
        self.date_from.setStyleSheet(DATE_STYLE)
        self.date_from.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.date_from)

        lbl_to = QLabel("To:")
        lbl_to.setStyleSheet(f"color: {NAVY}; background: transparent; border: none;")
        filter_layout.addWidget(lbl_to)

        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd/MM/yyyy")
        if max_date:
            self.date_to.setDate(QDate.fromString(max_date, "yyyy-MM-dd"))
        else:
            self.date_to.setDate(QDate.currentDate())
        self.date_to.setStyleSheet(DATE_STYLE)
        self.date_to.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self.date_to)

        layout.addWidget(filter_card)

        # ── Bulk action bar ──────────────────────────────────────────────
        bulk_bar = QFrame()
        bulk_bar.setObjectName("card")
        bulk_bar.setStyleSheet(CARD_STYLE)
        bulk_layout = QHBoxLayout(bulk_bar)
        bulk_layout.setContentsMargins(16, 8, 16, 8)
        bulk_layout.setSpacing(12)

        hint = QLabel("Select rows (Ctrl/Shift+click) then bulk-assign, or double-click a single category")
        hint.setFont(font_body(9))
        hint.setStyleSheet(f"color: #9CA3AF; background: transparent; border: none;")
        bulk_layout.addWidget(hint, stretch=1)

        self.selection_label = QLabel("")
        self.selection_label.setFont(font_body(9))
        self.selection_label.setStyleSheet(f"color: {NAVY}; background: transparent; border: none; font-weight: 600;")
        bulk_layout.addWidget(self.selection_label)

        bulk_layout.addWidget(QLabel("Set to:").tap(lambda l: l.setStyleSheet(f"color: {NAVY}; background: transparent; border: none;")) if False else self._make_label("Set to:"))

        self.bulk_cat_combo = QComboBox()
        for group_name, cats in get_groups():
            for cat in cats:
                self.bulk_cat_combo.addItem(cat.display_name, cat.id)
        self.bulk_cat_combo.setMinimumWidth(160)
        self.bulk_cat_combo.setStyleSheet(COMBO_STYLE)
        bulk_layout.addWidget(self.bulk_cat_combo)

        self.bulk_apply_btn = QPushButton("Apply")
        self.bulk_apply_btn.setFixedHeight(32)
        self.bulk_apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.bulk_apply_btn.setEnabled(False)
        self.bulk_apply_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {NAVY};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 4px 16px;
                font-size: 10pt;
                font-weight: 600;
            }}
            QPushButton:hover {{ background-color: #2B4A8C; }}
            QPushButton:disabled {{ background-color: #9CA3AF; }}
        """)
        self.bulk_apply_btn.clicked.connect(self._bulk_apply_category)
        bulk_layout.addWidget(self.bulk_apply_btn)

        layout.addWidget(bulk_bar)

        # Table
        table_card = QFrame()
        table_card.setObjectName("card")
        table_card.setStyleSheet(CARD_STYLE)
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(2, 2, 2, 2)

        self.model = TransactionTableModel()
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setSortingEnabled(False)

        # Update selection label when selection changes
        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)

        # Category delegate for inline editing
        self.cat_delegate = CategoryDelegate(self.db, self.table)
        self.table.setItemDelegateForColumn(3, self.cat_delegate)

        # Column widths
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 100)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 180)

        self.table.setStyleSheet(TABLE_STYLE)

        table_layout.addWidget(self.table)

        # Bottom bar: count + export button
        bottom_row = QHBoxLayout()
        bottom_row.setContentsMargins(12, 4, 12, 4)

        self.count_label = QLabel()
        self.count_label.setStyleSheet(f"color: {TEXT_SECONDARY}; font-size: 9pt; background: white; border: none;")
        bottom_row.addWidget(self.count_label)

        bottom_row.addStretch()

        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setFixedHeight(30)
        self.export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #16A34A; color: white; border: none;
                border-radius: 6px; padding: 4px 16px; font-size: 9pt; font-weight: 600;
            }}
            QPushButton:hover {{ background-color: #15803D; }}
            QPushButton:disabled {{ background-color: #9CA3AF; }}
        """)
        self.export_btn.clicked.connect(self._export_excel)
        bottom_row.addWidget(self.export_btn)

        table_layout.addLayout(bottom_row)

        layout.addWidget(table_card, stretch=1)

        # Ctrl+E shortcut for export
        self._export_shortcut = QShortcut(QKeySequence("Ctrl+E"), self)
        self._export_shortcut.activated.connect(self._export_excel)

    def _make_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(f"color: {NAVY}; background: transparent; border: none;")
        return lbl

    def _on_selection_changed(self):
        rows = self.table.selectionModel().selectedRows()
        n = len(rows)
        if n > 1:
            self.selection_label.setText(f"{n} selected")
            self.bulk_apply_btn.setEnabled(True)
        elif n == 1:
            self.selection_label.setText("1 selected")
            self.bulk_apply_btn.setEnabled(True)
        else:
            self.selection_label.setText("")
            self.bulk_apply_btn.setEnabled(False)

    def _bulk_apply_category(self):
        new_cat_id = self.bulk_cat_combo.currentData()
        if not new_cat_id:
            return

        rows = self.table.selectionModel().selectedRows()
        if not rows:
            return

        updates = []
        for idx in rows:
            txn = self.model.get_transaction(idx.row())
            if txn:
                updates.append((new_cat_id, txn.id))
                txn.category = new_cat_id

        if updates:
            self.db.bulk_update_categories(updates)
            # Refresh entire model display
            self.model.beginResetModel()
            self.model.endResetModel()
            self.selection_label.setText(f"Updated {len(updates)} transactions")

    def refresh(self):
        self._load_data()

    def _on_filter_changed(self):
        self._load_data()

    def _load_data(self):
        start = self.date_from.date().toString("yyyy-MM-dd")
        end = self.date_to.date().toString("yyyy-MM-dd")
        cat = self.cat_combo.currentData()

        txns = self.db.get_transactions(start=start, end=end,
                                        category=cat if cat else None,
                                        limit=2000)

        # Client-side search — matches description, date (YYYY-MM-DD), or category
        search = self.search_input.text().strip().lower()
        if search:
            txns = [t for t in txns
                    if search in t.description.lower()
                    or search in t.trans_date.strftime("%Y-%m-%d")
                    or search in t.trans_date.strftime("%d/%m/%Y")
                    or search in (t.category or "").lower()]

        self.model.set_data(txns)
        self.count_label.setText(f"{len(txns)} transaction{'s' if len(txns) != 1 else ''}")
        self.export_btn.setEnabled(len(txns) > 0)

    def _export_excel(self):
        """Export currently displayed transactions to an Excel file using openpyxl."""
        txns = self.model._data
        if not txns:
            QMessageBox.information(self, "Nothing to Export",
                                    "No transactions to export. Adjust your filters first.")
            return

        # Suggest a filename based on the date range
        start = self.date_from.date().toString("yyyy-MM")
        end = self.date_to.date().toString("yyyy-MM")
        suggested = f"BudgetBuddy_Transactions_{start}_to_{end}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Transactions", suggested,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return  # Cancelled

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

            wb = Workbook()
            ws = wb.active
            ws.title = "Transactions"

            # ── Header row ───────────────────────────────────────────────
            headers = ["Date", "Description", "Amount", "Category"]
            header_font = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
            header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            thin_border = Border(
                bottom=Side(style="thin", color="E5E7EB"),
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # ── Data rows ────────────────────────────────────────────────
            body_font = Font(name="Segoe UI", size=10)
            amount_fmt = '#,##0.00'

            for row_idx, txn in enumerate(txns, 2):
                # Date
                ws.cell(row=row_idx, column=1, value=txn.trans_date).font = body_font

                # Description
                ws.cell(row=row_idx, column=2, value=txn.description).font = body_font

                # Amount (as number, formatted)
                amount_cell = ws.cell(row=row_idx, column=3, value=txn.amount)
                amount_cell.font = body_font
                amount_cell.number_format = amount_fmt
                amount_cell.alignment = Alignment(horizontal="right")

                # Category (display name)
                cat_info = CATEGORIES.get(txn.category)
                cat_name = cat_info.display_name if cat_info else (txn.category or "Uncategorised").replace("_", " ").title()
                ws.cell(row=row_idx, column=4, value=cat_name).font = body_font

                # Subtle row border
                for c in range(1, 5):
                    ws.cell(row=row_idx, column=c).border = thin_border

            # ── Column widths ────────────────────────────────────────────
            ws.column_dimensions["A"].width = 14   # Date
            ws.column_dimensions["B"].width = 55   # Description
            ws.column_dimensions["C"].width = 14   # Amount
            ws.column_dimensions["D"].width = 20   # Category

            # ── Summary row ──────────────────────────────────────────────
            summary_row = len(txns) + 3
            ws.cell(row=summary_row, column=2, value="Total:").font = Font(name="Segoe UI", bold=True, size=10)
            total_cell = ws.cell(row=summary_row, column=3)
            total_cell.value = sum(t.amount for t in txns)
            total_cell.font = Font(name="Segoe UI", bold=True, size=10)
            total_cell.number_format = amount_fmt
            total_cell.alignment = Alignment(horizontal="right")

            ws.cell(row=summary_row, column=4,
                    value=f"{len(txns)} transactions").font = Font(name="Segoe UI", size=9, color="6B7280")

            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-filter
            ws.auto_filter.ref = f"A1:D{len(txns) + 1}"

            wb.save(file_path)

            QMessageBox.information(self, "Export Complete",
                                    f"Exported {len(txns)} transactions to:\n\n{file_path}")

        except ImportError:
            QMessageBox.warning(self, "Missing Package",
                                "openpyxl is required for Excel export.\n\n"
                                "Install with: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Could not export:\n\n{e}")
